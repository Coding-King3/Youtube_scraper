from Reddit_scraper import chromedriver_options_headers
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
import logging

import time
from openpyxl import Workbook, load_workbook
import os


def create_workbook():
    # This function creates an Excel workbook and check the condition whether the file is present or not,
    # if file is not present it will create it or else it will append it.

    excel_file_name = "../../Users/Bhavy/Downloads/Youtube_scraper_Data.xlsx"
    file_path = os.getcwd() + '\\' + excel_file_name
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        sheet = wb.active

        # Below three lines are written due to in some circumstances if the user deletes the headers in the file and then runs the code, This logic will place the headers again.

        if sheet['A1'].value is None:
            sheet.append(['Video Tittle', 'Channel Name', 'Views', 'Upload Date ', 'Video Url'])
            sheet.delete_rows(1, 1)
            wb.save(file_path)
    else:
        wb = Workbook()  # Creates workbook
        sheet = wb.active  # Makes workbook active
        sheet.title = 'Data'  # Renames sheet to Data
        headers = ['Video Tittle', 'Channel Name', 'Views', 'Upload Date ', 'Video Url']
        sheet.append(headers)
        wb.save(excel_file_name)
    return sheet, wb, excel_file_name


def finding_elements():
    try:

        # Visiting website and entering query getting video content

        driver.get('https://www.youtube.com/')
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(((By.
                                                                         XPATH,
                                                                         "//*[@class = 'yt-simple-endpoint style-scope ytd-mini-guide-entry-renderer']//span[contains(text(), 'History')]"))))
        time.sleep(1.5)
        driver.find_element(By.CSS_SELECTOR, "[placeholder = 'Search']").send_keys(search_keyword)
        time.sleep(1.4)
        driver.find_element(By.CSS_SELECTOR, "[class = 'ytSearchboxComponentSearchButton']").click()
        time.sleep(1)

        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(@class, 'yt-spec-button-shape-next') and @aria-label = 'Search filters']")))
        time.sleep(2)

        driver.find_element(By.XPATH,
                            "//*[contains(@class, 'yt-spec-button-shape-next') and @aria-label = 'Search filters']").click()

        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "[title = 'Search for Video']")))

        driver.find_element(By.CSS_SELECTOR, "[title = 'Search for Video']").click()
        WebDriverWait(driver, 15).until(EC.presence_of_element_located(
            (By.CSS_SELECTOR, "[id = 'dismissible'][class = 'style-scope ytd-video-renderer']")))
        time.sleep(2)

        # Reducing Zoom value to -25 so that elements load faster

        driver.execute_script("document.body.style.zoom='25%'")
        time.sleep(1)

        try:
            products_to_scrape = video_count

            # Show the ending scroll position of the webpage used for scrolling
            last_height = driver.execute_script("return document.documentElement.scrollHeight")

            # This while loop has two features, First condition is to scroll till we get our video_count data and then break the loop.
            # However if the video count is less on webpage than provided in input, it will scrape all the content available on the webpage and change the input to available video count

            while True:
                video_length = driver.find_elements(By.XPATH,
                                                    "(//*[@id = 'dismissible'] [@class = 'style-scope ytd-video-renderer']//*[@class = 'yt-simple-endpoint inline-block style-scope ytd-thumbnail'][not(contains(@href, 'shorts'))])/../../..").__len__()

                time.sleep(2)

                if video_length >= products_to_scrape:
                    return products_to_scrape
                else:
                    driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
                    time.sleep(3)
                    new_height = driver.execute_script("return document.documentElement.scrollHeight")
                    if new_height == last_height:
                        logging.error(
                            f'Error occured -- {video_length} videos on available on webpage as per the given count of {products_to_scrape} videos')

                        products_to_scrape = video_length
                        return products_to_scrape
                    last_height = new_height

        except Exception as error:
            logging.error("Video scrolling error")
            logging.error(error, exc_info=True)

    except WebDriverException as main_page_error:
        logging.error("Main Page error")
        logging.error( main_page_error, exc_info=True)

    # All except commands are logged in Youtube_scraper.log file


def scraping_youtube_data():
    driver.execute_script("window.scrollTo(document.documentElement.scrollHeight,0);")
    time.sleep(0.5)
    # driver.execute_script("document.body.style.zoom='25%'")
    time.sleep(1)

    for count in range(1, count_of_videos + 1):

        # The logic here is to wait for the xpath as per the given count,
        logging.info("Current count ", count)

        try:
            WebDriverWait(driver, 15).until(EC.presence_of_element_located(
                (By.XPATH,
                 f"(//*[@id = 'dismissible'] [@class = 'style-scope ytd-video-renderer']//*[@class = 'yt-simple-endpoint inline-block style-scope ytd-thumbnail'][not(contains(@href, 'shorts'))])[{count}]/../../..")))

            video_div = driver.find_element(By.XPATH,
                                            f"(//*[@id = 'dismissible'] [@class = 'style-scope ytd-video-renderer']//*[@class = 'yt-simple-endpoint inline-block style-scope ytd-thumbnail'][not(contains(@href, 'shorts'))])[{count}]/../../..")

            time.sleep(1)

            # Scraping - (title, channel_name, views, upload_time, video_url) below

            try:
                views = video_div.find_element(By.CSS_SELECTOR,
                                               "[class = 'inline-metadata-item style-scope ytd-video-meta-block']:nth-of-type(1)").text
            except WebDriverException as view_error:
                logging.error('view_count not there')

                logging.error( view_error, exc_info=True)
                views = None

            time.sleep(0.5)

            try:

                upload_time = video_div.find_element(By.CSS_SELECTOR,
                                                     "[class = 'inline-metadata-item style-scope ytd-video-meta-block']:nth-of-type(2)").text
            except WebDriverException as upload_time_error:
                upload_time = None
                logging.error('upload time not there')
                logging.error( upload_time_error, exc_info=True)

            time.sleep(0.5)

            try:

                title = video_div.find_element(By.CSS_SELECTOR, "[class = 'style-scope ytd-video-renderer'] h3").text

            except WebDriverException as title_error:
                title = None
                logging.error("title not there")
                logging.error(title_error, exc_info=True)
            time.sleep(0.5)

            try:

                channel_name = video_div.find_element(By.CSS_SELECTOR, "[id = 'channel-info']").text

            except WebDriverException as channel_name_error:
                channel_name = None
                logging.error("Channel name error")
                logging.error(channel_name_error, exc_info=True)

            time.sleep(0.5)

            try:
                video_url = video_div.find_element(By.CSS_SELECTOR,
                                                   "[id = 'dismissible'] [class = 'style-scope ytd-video-renderer'] a:nth-of-type(1)").get_attribute(
                    'href')
            except WebDriverException as video_url_error:
                video_url = None
                logging.error("Video url error")
                logging.error(video_url_error, exc_info=True)

            # Once the data is scraped, it's converted to a tuple and then appended to the sheet.
            # Once appended, it will be saved, so even in any scenario if the code stops, The data scraped till now will be saved.
            print(count)

            gather_data = (title, channel_name, views, upload_time, video_url)
            add_data.append(gather_data)
            save_data.save(file_name)

        except WebDriverException as scraping_data_error:
            logging.error("Problem loading in xpath")
            logging.error(scraping_data_error, exc_info=True)


if __name__ == '__main__':
    start_time = time.time()

    logging.basicConfig(filename='Youtube_scraper_.text', level=logging.ERROR,
                        format='%(asctime)s - %(levelname)s - %(message)s',
                        force = True)

    add_data, save_data, file_name = create_workbook()
    try:
        # Imported from Reddit_scrapper.py
        driver = chromedriver_options_headers()
        search_keyword = input("Enter your search query here ")
        video_count = int(input("Provide a count of videos that you need to scrape data for "))

        count_of_videos = finding_elements()
        scraping_youtube_data()
        driver.close()
    except Exception as main_error:
        logging.error(main_error, exc_info=True)

    finally:
        end_time = time.time()
        seconds = end_time - start_time
        convert_minutes = seconds / 60
        print(f"Program took {convert_minutes:.4f} minutes to run.")

